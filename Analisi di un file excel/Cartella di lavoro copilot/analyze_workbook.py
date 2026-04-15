from collections import Counter
from datetime import date, datetime, time
from decimal import Decimal

from openpyxl import load_workbook

WORKBOOK_PATH = r"S:\Progetti su github\corsi.ai\repo di riferimento\corsi.ai\Analisi di un file excel\Cartella di lavoro copilot\25.xlsx"


def cell_value_to_text(value):
    if value is None:
        return "<EMPTY>"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return repr(value)


def row_to_text(values):
    return "[" + ", ".join(cell_value_to_text(v) for v in values) + "]"


def is_empty(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def normalize_header(value, index):
    if is_empty(value):
        return f"<EMPTY_HEADER_{index}>"
    return str(value)


def style_summary(cell):
    details = []
    if cell.font:
        if cell.font.bold:
            details.append("bold")
        if cell.font.italic:
            details.append("italic")
        if (
            cell.font.color
            and getattr(cell.font.color, "type", None) == "rgb"
            and cell.font.color.rgb not in (None, "00000000", "FF000000")
        ):
            details.append(f"font_color={cell.font.color.rgb}")
    if cell.fill and cell.fill.fill_type:
        fg = getattr(cell.fill.fgColor, "rgb", None)
        details.append(f"fill={cell.fill.fill_type}:{fg}")
    if cell.number_format and cell.number_format != "General":
        details.append(f"number_format={cell.number_format}")
    if cell.alignment:
        if cell.alignment.horizontal:
            details.append(f"h_align={cell.alignment.horizontal}")
        if cell.alignment.vertical:
            details.append(f"v_align={cell.alignment.vertical}")
        if cell.alignment.text_rotation:
            details.append(f"text_rotation={cell.alignment.text_rotation}")
        if cell.alignment.wrap_text:
            details.append("wrap_text")
    if cell.protection:
        if cell.protection.locked is False:
            details.append("unlocked")
        if cell.protection.hidden:
            details.append("formula_hidden")
    if details:
        return ", ".join(details)
    return f"style_id={cell.style_id} (no extra style details)"


def detect_column_kind(values):
    nonempty = [v for v in values if not is_empty(v)]
    if not nonempty:
        return "empty"
    numeric = 0
    dates = 0
    text = 0
    booleans = 0
    formulas = 0
    for v in nonempty:
        if isinstance(v, bool):
            booleans += 1
        elif isinstance(v, (datetime, date, time)):
            dates += 1
        elif isinstance(v, (int, float, Decimal)):
            numeric += 1
        elif isinstance(v, str) and v.startswith("="):
            formulas += 1
        else:
            text += 1
    total = len(nonempty)
    unique_count = len({str(v) for v in nonempty})
    if dates == total:
        return f"date (nonempty={total}, unique={unique_count})"
    if numeric + formulas == total and numeric > 0:
        return f"numeric (nonempty={total}, unique={unique_count}, formulas={formulas})"
    if text == total:
        if unique_count <= max(20, int(total * 0.5)):
            return f"categorical (nonempty={total}, unique={unique_count})"
        return f"text (nonempty={total}, unique={unique_count})"
    if text > 0 and unique_count <= max(20, int(total * 0.5)):
        return (
            "categorical/mixed "
            f"(nonempty={total}, unique={unique_count}, numeric={numeric}, "
            f"date={dates}, bool={booleans}, formulas={formulas})"
        )
    return (
        "mixed "
        f"(nonempty={total}, unique={unique_count}, numeric={numeric}, "
        f"date={dates}, text={text}, bool={booleans}, formulas={formulas})"
    )


def detect_aggregation_row(values):
    lowered = [str(v).strip().lower() for v in values if not is_empty(v)]
    keywords = ["total", "subtotal", "grand total", "sum", "totale"]
    if any(any(keyword in item for keyword in keywords) for item in lowered):
        return True
    formula_hits = 0
    for v in values:
        if isinstance(v, str) and v.startswith("="):
            upper_value = v.upper()
            if any(fn in upper_value for fn in ["SUM(", "SUBTOTAL(", "AVERAGE(", "COUNT("]):
                formula_hits += 1
    return formula_hits > 0


wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
print(f"WORKBOOK: {WORKBOOK_PATH}")
print(f"SHEET_COUNT: {len(wb.sheetnames)}")
print("SHEET_NAMES:")
for idx, name in enumerate(wb.sheetnames, start=1):
    print(f"  {idx}. {name}")
print()

for sheet_index, ws in enumerate(wb.worksheets, start=1):
    print("=" * 100)
    print(f"SHEET {sheet_index}: {ws.title}")
    print(f"DIMENSIONS: rows={ws.max_row}, cols={ws.max_column}")
    print(f"FREEZE_PANES: {ws.freeze_panes if ws.freeze_panes else 'None'}")
    auto_filter_ref = ws.auto_filter.ref if ws.auto_filter and ws.auto_filter.ref else None
    print(f"AUTOFILTER: {auto_filter_ref if auto_filter_ref else 'None'}")

    merged = [str(rng) for rng in ws.merged_cells.ranges]
    print(f"MERGED_CELLS_COUNT: {len(merged)}")
    if merged:
        for item in merged:
            print(f"  MERGED: {item}")
    else:
        print("  MERGED: None")

    print(f"TABLE_COUNT: {len(ws.tables)}")
    if ws.tables:
        for table_name, table in ws.tables.items():
            headers = [col.name for col in table.tableColumns]
            print(
                f"  TABLE: name={table_name}, ref={table.ref}, "
                f"totalsRowShown={table.totalsRowShown}, headers={headers}"
            )
    else:
        print("  TABLE: None")

    header_values = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
    print(f"HEADERS_ROW_1: {row_to_text(header_values)}")

    nonempty_cells = 0
    actual_data_rows = []
    empty_cells = []
    aggregation_rows = []
    style_counter = Counter()
    style_cells = []
    hidden_rows = []
    hidden_cols = []

    for row_idx in range(1, ws.max_row + 1):
        row_values = []
        row_has_data = False
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            row_values.append(value)
            if not is_empty(value):
                nonempty_cells += 1
                row_has_data = True
                style_counter[cell.style_id] += 1
                style_cells.append(cell)
            else:
                empty_cells.append(cell.coordinate)
        if row_idx > 1 and row_has_data and len(actual_data_rows) < 10:
            actual_data_rows.append((row_idx, row_values))
        if row_idx > 1 and row_has_data and detect_aggregation_row(row_values):
            aggregation_rows.append((row_idx, row_values))
        if ws.row_dimensions[row_idx].hidden:
            hidden_rows.append(row_idx)

    for col_idx in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        if ws.column_dimensions[col_letter].hidden:
            hidden_cols.append(col_letter)

    total_cells = ws.max_row * ws.max_column if ws.max_row and ws.max_column else 0
    empty_ratio = 1.0 if total_cells == 0 else len(empty_cells) / total_cells
    print(f"NONEMPTY_CELL_COUNT: {nonempty_cells}")
    print(f"EMPTY_CELL_COUNT_IN_USED_RANGE: {len(empty_cells)}")
    print(f"EMPTY_RATIO_IN_USED_RANGE: {empty_ratio:.4f}")
    if nonempty_cells == 0:
        print("SHEET_EMPTY_STATUS: empty sheet")
    elif empty_ratio >= 0.9:
        print("SHEET_EMPTY_STATUS: mostly-empty sheet")
    else:
        print("SHEET_EMPTY_STATUS: populated sheet")

    print("SAMPLE_DATA_ROWS (up to 10 actual data rows):")
    if actual_data_rows:
        for row_idx, values in actual_data_rows:
            print(f"  ROW {row_idx}: {row_to_text(values)}")
    else:
        print("  None")

    print("COLUMN_TYPE_ANALYSIS:")
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=1, column=col_idx).value, col_idx)
        values = [ws.cell(row=row_idx, column=col_idx).value for row_idx in range(2, ws.max_row + 1)]
        print(f"  COLUMN {col_idx} ({header}): {detect_column_kind(values)}")

    print("ANOMALIES:")
    if empty_cells:
        print("  EMPTY_CELLS:")
        for coord in empty_cells:
            print(f"    {coord}")
    else:
        print("  EMPTY_CELLS: None")

    if aggregation_rows:
        print("  AGGREGATION_ROWS:")
        for row_idx, values in aggregation_rows:
            print(f"    ROW {row_idx}: {row_to_text(values)}")
    else:
        print("  AGGREGATION_ROWS: None")

    if hidden_rows or hidden_cols:
        print("  HIDDEN_ELEMENTS:")
        if hidden_rows:
            print(f"    Hidden rows: {hidden_rows}")
        if hidden_cols:
            print(f"    Hidden columns: {hidden_cols}")
    else:
        print("  HIDDEN_ELEMENTS: None")

    baseline_style = style_counter.most_common(1)[0][0] if style_counter else None
    unusual_style_cells = []
    if baseline_style is not None:
        for cell in style_cells:
            if cell.style_id != baseline_style:
                unusual_style_cells.append(cell)

    if unusual_style_cells:
        print("  UNUSUAL_FORMATTING:")
        for cell in unusual_style_cells:
            print(
                f"    {cell.coordinate}: value={cell_value_to_text(cell.value)}; "
                f"style_id={cell.style_id}; {style_summary(cell)}"
            )
    else:
        print("  UNUSUAL_FORMATTING: None")

    print()
